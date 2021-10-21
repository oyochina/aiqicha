
import caiji
from openpyxl import Workbook,load_workbook
import time,re
book= load_workbook('汇龙最低2422.xlsx')

# 获取sheet页
ws = book['Sheet1']


pattern = re.compile(r'<[^>]+>', re.S)

j=0

for i in range(401,500):
    name,address = caiji.collection(ws.cell(row=i, column=2).value)
    
    result = pattern.sub('', address)
    ws.cell(i,12).value=result
    j=j+1
    if j==10:
        book.save('汇龙最低2422.xlsx')
        j=0



    print(str(i)+name+':'+result)
    
    time.sleep(2)
    
    #print(ws.cell(row=i, column=3).value)






#r=requests.get(url,params=kv)
#r.encoding='utf-8'
#print(r.request.url)


#path='test.xlsx'
#wb=Workbook()
#ws=wb.active
#ws.title='hello'
#ws.cell(row=1,column=2,value=item['entName'])

#entName=item['entName']
#pattern = re.compile(r'<[^>]+>', re.S)
#result = pattern.sub('', entName)
#ws.cell(2,6).value=result

book.close